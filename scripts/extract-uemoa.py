# -*- coding: utf-8 -*-
"""Extrait PROJET UEMOA.xlsx vers les modules de donnees du portfolio.

    python scripts/extract-uemoa.py ["chemin/vers/PROJET UEMOA.xlsx"]

Le classeur source n'est pas versionne : il reste chez son auteur, et seules
les valeurs extraites sont publiees. Sans argument, le script le cherche dans
le dossier Telechargements de l'utilisateur courant.

Dependance : openpyxl (pip install openpyxl).
"""
import json, re, sys
from pathlib import Path

import openpyxl

DEFAULT_SRC = Path.home() / "Downloads" / "PROJET DATA" / "PROJET UEMOA.xlsx"
SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
if not SRC.exists():
    sys.exit(f"Classeur introuvable : {SRC}\nPassez son chemin en argument.")

LIB = Path(__file__).resolve().parent.parent / "lib"
OUT_DASHBOARD = LIB / "uemoa-dashboard.ts"
OUT_SHEETS = LIB / "uemoa-sheets.ts"

NBSP = "\u00a0"

wbv = openpyxl.load_workbook(SRC, data_only=True)
wbf = openpyxl.load_workbook(SRC)


def decimals(fmt):
    m = re.search(r"\.(0+)", fmt)
    return len(m.group(1)) if m else 0


def fr(x, dec):
    s = f"{x:,.{dec}f}".replace(",", NBSP).replace(".", ",")
    return s


def fmt_cell(value, number_format):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "VRAI" if value else "FAUX"
    if isinstance(value, str):
        return value
    if not isinstance(value, (int, float)):
        return str(value)

    fmt = (number_format or "General").split(";")[0]
    if "%" in fmt:
        return fr(value * 100, decimals(fmt)) + f"{NBSP}%"
    if "#,##0" in fmt:
        return fr(value, decimals(fmt))
    if re.search(r"\.0+", fmt):
        return fr(value, decimals(fmt))
    # General: gros entiers séparés, décimales coupées à deux. Le seuil est à
    # 10 000 et non 1 000 pour que les millésimes (2020…2024), stockés en
    # General, ne ressortent pas en « 2 020 ».
    if float(value).is_integer() and abs(value) >= 10000:
        return fr(value, 0)
    if float(value).is_integer():
        return str(int(value))
    return fr(value, 2)


def grid(name, max_row=None, max_col=None):
    wsv, wsf = wbv[name], wbf[name]
    rows = []
    mr = max_row or wsv.max_row
    mc = max_col or wsv.max_column
    for r in range(1, mr + 1):
        row = []
        for c in range(1, mc + 1):
            cv, cf = wsv.cell(r, c), wsf.cell(r, c)
            text = fmt_cell(cv.value, cf.number_format)
            cell = {"v": text}
            if isinstance(cv.value, (int, float)) and not isinstance(cv.value, bool):
                cell["n"] = True
            if cf.font and cf.font.bold:
                cell["b"] = True
            row.append(cell)
        rows.append(row)
    # openpyxl compte parfois une ligne vide en fin de feuille : on la retire
    # pour ne pas afficher une ligne fantôme sous le tableau.
    while rows and all(c["v"] == "" for c in rows[-1]):
        rows.pop()
    return {"name": name, "cols": mc, "rows": rows}


# --- Dashboard: read as structured content rather than as a raw grid, so the
# --- page can lay it out properly instead of imitating merged Excel cells.
db = wbv["Tableau de bord"]


def txt(coord):
    v = db[coord].value
    return "" if v is None else str(v)


kpis = []
for row_label, row_value, row_note in ((4, 5, 6), (8, 9, 10)):
    for col in ("A", "C", "E", "G"):
        kpis.append({
            "label": txt(f"{col}{row_label}"),
            "value": txt(f"{col}{row_value}"),
            "note": txt(f"{col}{row_note}"),
        })

ranking = []
for r in range(18, 26):
    ranking.append({
        "rank": int(db[f"A{r}"].value),
        "pays": txt(f"B{r}"),
        "taux": round(float(db[f"C{r}"].value), 2),
        "ouverts": int(db[f"D{r}"].value),
        "actifs": int(db[f"E{r}"].value),
    })

dashboard = {
    "title": txt("A1"),
    "subtitle": txt("A2"),
    "kpis": kpis,
    "answerTitle": txt("A12"),
    "answer": txt("A13"),
    "rankingTitle": txt("A16"),
    "ranking": ranking,
}

# --- Chart series, read from the ranges the workbook's own charts point at.
u = wbv["UEMOA"]
comptes, taux_zone, indices = [], [], []
for r in range(3, 8):
    annee = str(u[f"AG{r}"].value)
    comptes.append({
        "annee": annee,
        "ouverts": int(u[f"AH{r}"].value),
        "actifs": int(u[f"AI{r}"].value),
    })
    taux_zone.append({"annee": annee, "taux": round(float(u[f"AL{r}"].value), 2)})
    # Deux décimales : les étiquettes de fin de courbe en affichent deux, et
    # arrondir à une seule donnait « 264,00 » pour un indice réel de 263,95.
    indices.append({
        "annee": annee,
        "ouverts": round(float(u[f"AA{r}"].value), 2),
        "actifs": round(float(u[f"AB{r}"].value), 2),
    })

p = wbv["PAYS"]
taux_pays = []
for r in range(3, 11):
    taux_pays.append({
        "pays": str(p[f"AE{r}"].value),
        "taux": round(float(p[f"AF{r}"].value), 2),
    })

charts = {
    "comptes": comptes,
    "tauxZone": taux_zone,
    "indices": indices,
    "tauxPays": taux_pays,
}

sheets = [grid("Analyse"), grid("UEMOA"), grid("PAYS"), grid("TCD_Pays")]


def js(obj):
    return json.dumps(obj, ensure_ascii=False, indent=2)


NOTE = '''/**
 * Contenu de PROJET UEMOA.xlsx, extrait via scripts/extract-uemoa.py.
 *
 * Le classeur est rendu en HTML plutôt que servi en téléchargement : le
 * visiteur voit les mêmes feuilles, les mêmes valeurs et les mêmes graphiques,
 * sans qu'aucun fichier modifiable ne quitte le site.
 *
 * NE PAS ÉDITER À LA MAIN — relancer le script après toute modification du
 * classeur source.
 */
'''

SHEET_TYPES = '''
export type Cell = {
  /** Valeur déjà formatée pour l'affichage (séparateurs et décimales fr-FR). */
  v: string;
  /** Nombre : aligné à droite dans la grille. */
  n?: boolean;
  /** Gras dans le classeur d'origine : titres de blocs et en-têtes. */
  b?: boolean;
};

export type SheetGrid = {
  name: string;
  cols: number;
  rows: Cell[][];
};

'''

# Deux modules plutôt qu'un seul : la page d'accueil n'affiche que les quatre
# indicateurs de tête et n'a aucune raison d'embarquer le millier de cellules
# des feuilles avec eux — un seul module les mettait dans le même chunk.
with open(OUT_DASHBOARD, "w", encoding="utf-8") as f:
    f.write(NOTE + "\n")
    f.write("export const dashboard = " + js(dashboard) + " as const;\n\n")
    f.write("export const charts = " + js(charts) + " as const;\n")

with open(OUT_SHEETS, "w", encoding="utf-8") as f:
    f.write(NOTE)
    f.write(SHEET_TYPES)
    f.write("export const sheetGrids: SheetGrid[] = " + js(sheets) + ";\n")

print("written:", OUT_DASHBOARD, "+", OUT_SHEETS)
print("sheets:", [(s["name"], len(s["rows"]), s["cols"]) for s in sheets])
print("kpis:", len(kpis), "ranking:", len(ranking))
